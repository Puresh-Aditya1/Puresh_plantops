from fastapi import APIRouter, Depends, HTTPException
from datetime import datetime, timezone, timedelta
from typing import Optional
import io
import json
from bson import ObjectId
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse
import zipfile

from database import db
from auth import get_current_user

router = APIRouter(prefix="/api")


def require_admin(current_user: dict = Depends(get_current_user)):
    """Dependency to require admin role"""
    if current_user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Collections to archive
ARCHIVE_COLLECTIONS = [
    "batches",
    "packing_entries", 
    "dispatches",
    "finished_product_receives",
    "finished_product_repacks",
    "finished_product_wastages",
    "milk_entries",
    "raw_material_stocks",
    "raw_material_adjustments",
    "raw_material_consumptions"
]


def serialize_doc(doc):
    """Convert MongoDB document to JSON-serializable dict"""
    if doc is None:
        return None
    result = {}
    for key, value in doc.items():
        if key == '_id':
            continue
        elif isinstance(value, ObjectId):
            result[key] = str(value)
        elif isinstance(value, datetime):
            result[key] = value.isoformat()
        elif isinstance(value, list):
            result[key] = [serialize_doc(v) if isinstance(v, dict) else v for v in value]
        elif isinstance(value, dict):
            result[key] = serialize_doc(value)
        else:
            result[key] = value
    return result


@router.get("/archive/preview")
async def preview_archive(
    archive_date: str,
    current_user: dict = Depends(require_admin)
):
    """Preview what will be archived for a given date"""
    try:
        # Parse and validate date
        cutoff_date = datetime.strptime(archive_date, "%Y-%m-%d").date()
        cutoff_str = archive_date
        
        preview = {}
        
        # Count batches
        preview["batches"] = await db.batches.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count packing entries (from finished_products)
        preview["packing_entries"] = await db.finished_products.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count dispatches
        preview["dispatches"] = await db.dispatches.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count receives
        preview["finished_product_receives"] = await db.finished_product_receives.count_documents({
            "receive_date": {"$lte": cutoff_str}
        })
        
        # Count repacks
        preview["finished_product_repacks"] = await db.finished_product_repacks.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count wastages
        preview["finished_product_wastages"] = await db.finished_product_wastages.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count milk entries (from milk_stock collection)
        preview["milk_entries"] = await db.milk_stock.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count raw material stocks (purchases)
        preview["raw_material_stocks"] = await db.raw_material_stock.count_documents({
            "date": {"$lte": cutoff_str}
        })
        
        # Count raw material adjustments
        preview["raw_material_adjustments"] = await db.raw_material_adjustments.count_documents({
            "adjustment_date": {"$lte": cutoff_str}
        })
        
        # Count raw material consumptions
        preview["raw_material_consumptions"] = await db.raw_material_consumptions.count_documents({
            "consumption_date": {"$lte": cutoff_str}
        })
        
        total_records = sum(preview.values())
        
        return {
            "archive_date": archive_date,
            "collections": preview,
            "total_records": total_records
        }
        
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")


@router.get("/archive/closing-stock")
async def get_closing_stock(
    archive_date: str,
    current_user: dict = Depends(require_admin)
):
    """Calculate closing stock as of archive date"""
    cutoff_str = archive_date
    
    # Calculate semi-finished closing stock
    semi_finished_stock = {}
    
    # From batches (production)
    batches = await db.batches.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for batch in batches:
        if batch.get("output_type") == "semi-finished":
            name = batch.get("product_name", "Unknown")
            qty = float(batch.get("quantity_produced", 0))
            semi_finished_stock[name] = semi_finished_stock.get(name, 0) + qty
    
    # From packing/finished_products (consumption of semi-finished)
    packings = await db.finished_products.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for pk in packings:
        # Get semi-finished product name from the linked record
        sf_id = pk.get("semi_finished_id")
        if sf_id:
            sf = await db.semi_finished_products.find_one({"id": sf_id}, {"_id": 0, "product_name": 1})
            name = sf.get("product_name", "Unknown") if sf else "Unknown"
        else:
            name = "Unknown"
        consumed = float(pk.get("semi_finished_consumed", 0))
        semi_finished_stock[name] = semi_finished_stock.get(name, 0) - consumed
    
    # Calculate finished product closing stock
    finished_stock = {}
    
    # From batches (direct finished)
    for batch in batches:
        if batch.get("output_type") == "finished":
            sku = batch.get("product_name", "Unknown")
            qty = float(batch.get("quantity_produced", 0))
            finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From packing/finished_products (production of finished goods from semi-finished only)
    # Skip source="receive" (counted in receives below) and source="batch" (counted in batches above)
    for pk in packings:
        source = pk.get("source")
        if source in ("receive", "batch"):
            continue
        sku = pk.get("sku", "Unknown")
        qty = float(pk.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From receives
    receives = await db.finished_product_receives.find({"receive_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for r in receives:
        sku = r.get("sku", "Unknown")
        qty = float(r.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From repacks
    repacks = await db.finished_product_repacks.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for rp in repacks:
        source = rp.get("source_sku", "Unknown")
        target = rp.get("target_sku", "Unknown")
        used = float(rp.get("quantity_used", 0))
        produced = float(rp.get("quantity_produced", 0))
        finished_stock[source] = finished_stock.get(source, 0) - used
        finished_stock[target] = finished_stock.get(target, 0) + produced
    
    # From dispatches
    dispatches = await db.dispatches.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for d in dispatches:
        for prod in d.get("products", []):
            sku = prod.get("sku", "Unknown")
            qty = float(prod.get("quantity", 0))
            finished_stock[sku] = finished_stock.get(sku, 0) - qty
    
    # From wastages
    wastages = await db.finished_product_wastages.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for w in wastages:
        sku = w.get("sku", "Unknown")
        qty = float(w.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) - qty
    
    # Calculate raw material closing stock
    raw_material_stock = {}
    
    # From initial stocks
    rm_initials = await db.initial_stocks.find({"type": "raw_material"}, {"_id": 0}).to_list(100)
    for init in rm_initials:
        name = init.get("name", "Unknown")
        qty = float(init.get("quantity", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) + qty
    
    # From daily stock entries (contains purchased and used)
    rm_stocks = await db.raw_material_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for rm in rm_stocks:
        name = rm.get("name", "Unknown")
        purchased = float(rm.get("purchased", 0))
        used = float(rm.get("used", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) + purchased - used
    
    # From adjustments
    rm_adjs = await db.rm_adjustments.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for adj in rm_adjs:
        name = adj.get("material_name", adj.get("name", "Unknown"))
        qty = float(adj.get("quantity", 0))
        adj_type = adj.get("adjustment_type", adj.get("type", "gain"))
        if adj_type == "loss":
            qty = -qty
        raw_material_stock[name] = raw_material_stock.get(name, 0) + qty
    
    # From direct consumptions (if any)
    rm_cons = await db.rm_direct_consumption.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for con in rm_cons:
        name = con.get("material_name", con.get("name", "Unknown"))
        qty = float(con.get("quantity", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) - qty
    
    # Calculate milk closing stock (including Fat and SNF)
    milk_qty = 0
    milk_fat = 0
    milk_snf = 0
    
    # From initial stocks
    milk_initials = await db.initial_stocks.find({"type": "milk"}, {"_id": 0}).to_list(100)
    for init in milk_initials:
        milk_qty += float(init.get("quantity", 0))
        milk_fat += float(init.get("fat_kg", 0))
        milk_snf += float(init.get("snf_kg", 0))
    
    # From milk_stock entries (received/purchased milk)
    milk_stock_entries = await db.milk_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for entry in milk_stock_entries:
        milk_qty += float(entry.get("quantity_kg", 0))
        milk_fat += float(entry.get("fat_kg", 0))
        milk_snf += float(entry.get("snf_kg", 0))
    
    # From batches (milk consumed in production)
    for batch in batches:
        batch_milk_kg = float(batch.get("milk_kg", 0))
        if batch_milk_kg > 0:
            milk_qty -= batch_milk_kg
            # Calculate fat and snf used based on batch percentages
            fat_pct = float(batch.get("fat_percent", 0))
            snf_pct = float(batch.get("snf_percent", 0))
            milk_fat -= (fat_pct / 100) * batch_milk_kg
            milk_snf -= (snf_pct / 100) * batch_milk_kg
    
    # From milk wastages
    milk_wastages = await db.milk_wastages.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for w in milk_wastages:
        milk_qty -= float(w.get("quantity_kg", 0))
        milk_fat -= float(w.get("fat_kg", 0))
        milk_snf -= float(w.get("snf_kg", 0))
    
    # From milk adjustments
    milk_adjs = await db.milk_adjustments.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(10000)
    for adj in milk_adjs:
        qty = float(adj.get("quantity_kg", 0))
        fat = float(adj.get("fat_kg", 0))
        snf = float(adj.get("snf_kg", 0))
        adj_type = adj.get("type", "gain")
        if adj_type == "loss":
            qty, fat, snf = -qty, -fat, -snf
        milk_qty += qty
        milk_fat += fat
        milk_snf += snf
    
    milk_stock = {
        "Milk": round(milk_qty, 2),
        "Fat": round(milk_fat, 2),
        "SNF": round(milk_snf, 2)
    }
    
    # Calculate the next day for opening balance
    archive_dt = datetime.strptime(archive_date, "%Y-%m-%d")
    next_day = (archive_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    
    # Build opening balances that will be created
    opening_balances = {
        "date": next_day,
        "semi_finished": {k: v for k, v in semi_finished_stock.items() if v > 0},
        "finished_products": {k: v for k, v in finished_stock.items() if v > 0},
        "raw_materials": {k: v for k, v in raw_material_stock.items() if v > 0},
        "milk": {
            "quantity_kg": round(milk_qty, 2) if milk_qty > 0 else 0,
            "fat_kg": round(milk_fat, 2) if milk_fat > 0 else 0,
            "snf_kg": round(milk_snf, 2) if milk_snf > 0 else 0
        } if milk_qty > 0 else {}
    }
    
    return {
        "archive_date": archive_date,
        "semi_finished_stock": semi_finished_stock,
        "finished_stock": finished_stock,
        "raw_material_stock": raw_material_stock,
        "milk_stock": milk_stock,
        "opening_balances_to_create": opening_balances
    }


def create_excel_sheet(wb, sheet_name, data, columns):
    """Create an Excel sheet with formatted data"""
    ws = wb.create_sheet(title=sheet_name)
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, item in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = item.get(col_name.lower().replace(" ", "_"), "")
            if isinstance(value, list):
                value = json.dumps(value)
            elif isinstance(value, dict):
                value = json.dumps(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    return ws


@router.post("/archive/execute")
async def execute_archive(
    archive_date: str,
    current_user: dict = Depends(require_admin)
):
    """Execute the archive operation"""
    try:
        cutoff_str = archive_date
        archive_id = f"archive_{archive_date}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
        
        # Get closing stock before archiving
        closing_stock = await get_closing_stock(archive_date, current_user)
        
        archived_counts = {}
        
        # Archive batches
        batches = await db.batches.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if batches:
            for b in batches:
                b["archive_id"] = archive_id
            await db.batches_archive.insert_many(batches)
            await db.batches.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["batches"] = len(batches)
        
        # Archive finished_products (packing entries) and DELETE old records
        finished_prods = await db.finished_products.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if finished_prods:
            for p in finished_prods:
                p["archive_id"] = archive_id
                # Add semi_finished_product name for display
                if p.get('semi_finished_id'):
                    sf = await db.semi_finished_products.find_one({"id": p['semi_finished_id']}, {"_id": 0, "product_name": 1})
                    p['semi_finished_product'] = sf.get('product_name', 'Unknown') if sf else 'Unknown'
                if p.get('source') == 'batch':
                    p['semi_finished_product'] = '(Direct from Batch)'
                p['finished_product_sku'] = p.get('sku', '')
                p['packing_date'] = p.get('date', '')
            await db.packing_entries_archive.insert_many(finished_prods)
            # DELETE old finished_products - opening balances will be created below
            await db.finished_products.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["packing_entries"] = len(finished_prods)
        
        # Archive semi_finished_products and DELETE old records
        semi_finished = await db.semi_finished_products.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if semi_finished:
            for sf in semi_finished:
                sf["archive_id"] = archive_id
            await db.semi_finished_products_archive.insert_many(semi_finished)
            # DELETE old semi_finished_products - opening balances will be created via batches
            await db.semi_finished_products.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["semi_finished_products"] = len(semi_finished)
        
        # Archive dispatches
        dispatches = await db.dispatches.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if dispatches:
            for d in dispatches:
                d["archive_id"] = archive_id
            await db.dispatches_archive.insert_many(dispatches)
            await db.dispatches.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["dispatches"] = len(dispatches)
        
        # Archive receives
        receives = await db.finished_product_receives.find({"receive_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if receives:
            for r in receives:
                r["archive_id"] = archive_id
            await db.finished_product_receives_archive.insert_many(receives)
            await db.finished_product_receives.delete_many({"receive_date": {"$lte": cutoff_str}})
            archived_counts["receives"] = len(receives)
        
        # Archive repacks
        repacks = await db.finished_product_repacks.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if repacks:
            for rp in repacks:
                rp["archive_id"] = archive_id
            await db.finished_product_repacks_archive.insert_many(repacks)
            await db.finished_product_repacks.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["repacks"] = len(repacks)
        
        # Archive wastages
        wastages = await db.finished_product_wastages.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if wastages:
            for w in wastages:
                w["archive_id"] = archive_id
            await db.finished_product_wastages_archive.insert_many(wastages)
            await db.finished_product_wastages.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["wastages"] = len(wastages)
        
        # Archive milk stock entries
        milk_entries = await db.milk_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if milk_entries:
            for m in milk_entries:
                m["archive_id"] = archive_id
            await db.milk_stock_archive.insert_many(milk_entries)
            await db.milk_stock.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["milk_entries"] = len(milk_entries)
        
        # Archive raw material stocks
        rm_stocks = await db.raw_material_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if rm_stocks:
            for rm in rm_stocks:
                rm["archive_id"] = archive_id
            await db.raw_material_stock_archive.insert_many(rm_stocks)
            await db.raw_material_stock.delete_many({"date": {"$lte": cutoff_str}})
            archived_counts["raw_material_stocks"] = len(rm_stocks)
        
        # Archive raw material adjustments
        rm_adjs = await db.raw_material_adjustments.find({"adjustment_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if rm_adjs:
            for adj in rm_adjs:
                adj["archive_id"] = archive_id
            await db.raw_material_adjustments_archive.insert_many(rm_adjs)
            await db.raw_material_adjustments.delete_many({"adjustment_date": {"$lte": cutoff_str}})
            archived_counts["raw_material_adjustments"] = len(rm_adjs)
        
        # Archive raw material consumptions
        rm_cons = await db.raw_material_consumptions.find({"consumption_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
        if rm_cons:
            for con in rm_cons:
                con["archive_id"] = archive_id
            await db.raw_material_consumptions_archive.insert_many(rm_cons)
            await db.raw_material_consumptions.delete_many({"consumption_date": {"$lte": cutoff_str}})
            archived_counts["raw_material_consumptions"] = len(rm_cons)
        
        # Calculate next day for opening balance
        next_day = (datetime.strptime(archive_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        
        # Create opening balance entries for semi-finished products
        for product_name, stock in closing_stock["semi_finished_stock"].items():
            if stock > 0:
                batch_id = f"OB-SF-{next_day}-{product_name}"
                # Create opening balance batch
                await db.batches.insert_one({
                    "id": batch_id,
                    "batch_number": f"OB-{next_day}",
                    "date": next_day,
                    "output_type": "semi-finished",
                    "product_name": product_name,
                    "quantity_produced": stock,
                    "milk_quantity": 0,
                    "milk_kg": 0,
                    "fat_percent": 0,
                    "snf_percent": 0,
                    "fat_rate": 0,
                    "snf_rate": 0,
                    "raw_materials": [],
                    "additional_costs": [],
                    "total_cost": 0,
                    "notes": f"Opening balance from archive {archive_date}",
                    "is_opening_balance": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                # Also create the semi_finished_products entry for stock tracking
                await db.semi_finished_products.insert_one({
                    "id": f"SF-OB-{next_day}-{product_name}",
                    "product_name": product_name,
                    "batch_id": batch_id,
                    "quantity_kg": stock,
                    "current_stock": stock,
                    "date": next_day,
                    "is_opening_balance": True,
                    "notes": f"Opening balance from archive {archive_date}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        
        # Create opening balance entries for finished products
        for sku, stock in closing_stock["finished_stock"].items():
            if stock > 0:
                receive_id = f"OB-FP-{next_day}-{sku}"
                await db.finished_product_receives.insert_one({
                    "id": receive_id,
                    "sku": sku,
                    "quantity": stock,
                    "receive_date": next_day,
                    "source_name": "Opening Balance",
                    "cost_per_unit": 0,
                    "notes": f"Opening balance from archive {archive_date}",
                    "is_opening_balance": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                # Also create the finished_products entry for stock tracking
                await db.finished_products.insert_one({
                    "id": f"FP-OB-{next_day}-{sku}",
                    "sku": sku,
                    "quantity": stock,
                    "quantity_wasted": 0,
                    "current_stock": stock,
                    "unit": "",
                    "source": "receive",
                    "source_receive_id": receive_id,
                    "date": next_day,
                    "is_opening_balance": True,
                    "notes": f"Opening balance from archive {archive_date}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        
        # Create opening balance entries for raw materials
        for material_name, stock in closing_stock["raw_material_stock"].items():
            if stock > 0:
                await db.raw_material_stock.insert_one({
                    "id": f"OB-RM-{next_day}-{material_name}",
                    "name": material_name,
                    "unit": "kg",
                    "date": next_day,
                    "opening_stock": 0,
                    "purchased": stock,  # Opening balance is treated as "purchased"
                    "used": 0,
                    "closing_stock": stock,
                    "cost_per_unit": 0,
                    "notes": f"Opening balance from archive {archive_date}",
                    "is_opening_balance": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        
        # Milk opening balance is NOT automatically created - user will add manually
        
        # Save archive metadata
        await db.archive_metadata.insert_one({
            "archive_id": archive_id,
            "archive_date": archive_date,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "executed_by": current_user["username"],
            "archived_counts": archived_counts,
            "closing_stock": closing_stock,
            "total_records": sum(archived_counts.values())
        })
        
        # Update settings with last archive date
        await db.settings.update_one(
            {"key": "last_archive_date"},
            {"$set": {"value": archive_date, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        
        return {
            "success": True,
            "archive_id": archive_id,
            "archive_date": archive_date,
            "archived_counts": archived_counts,
            "total_records": sum(archived_counts.values()),
            "opening_balances_created": True
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/archive/list")
async def list_archives(current_user: dict = Depends(require_admin)):
    """List all archives"""
    archives = await db.archive_metadata.find({}, {"_id": 0}).sort("executed_at", -1).to_list(100)
    return archives


@router.get("/archive/download/{archive_id}")
async def download_archive(
    archive_id: str,
    current_user: dict = Depends(require_admin)
):
    """Download archive as Excel files in a ZIP"""
    
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # Export batches
        batches = await db.batches_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if batches:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Batches"
            headers = ["ID", "Batch Number", "Date", "Output Type", "Product Name", "Quantity Produced", "Milk Quantity", "Total Cost", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, b in enumerate(batches, 2):
                ws.cell(row=row, column=1, value=b.get("id", ""))
                ws.cell(row=row, column=2, value=b.get("batch_number", ""))
                ws.cell(row=row, column=3, value=b.get("date", ""))
                ws.cell(row=row, column=4, value=b.get("output_type", ""))
                ws.cell(row=row, column=5, value=b.get("product_name", ""))
                ws.cell(row=row, column=6, value=b.get("quantity_produced", 0))
                ws.cell(row=row, column=7, value=b.get("milk_quantity", 0))
                ws.cell(row=row, column=8, value=b.get("total_cost", 0))
                ws.cell(row=row, column=9, value=b.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"batches_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export packing entries
        packings = await db.packing_entries_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if packings:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Packing Entries"
            headers = ["ID", "Packing Date", "Semi-Finished Product", "Finished SKU", "Quantity Produced", "Semi-Finished Consumed", "Total Cost", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, p in enumerate(packings, 2):
                ws.cell(row=row, column=1, value=p.get("id", ""))
                ws.cell(row=row, column=2, value=p.get("packing_date", ""))
                ws.cell(row=row, column=3, value=p.get("semi_finished_product", ""))
                ws.cell(row=row, column=4, value=p.get("finished_product_sku", ""))
                ws.cell(row=row, column=5, value=p.get("quantity_produced", 0))
                ws.cell(row=row, column=6, value=p.get("semi_finished_consumed", 0))
                ws.cell(row=row, column=7, value=p.get("total_packing_cost", 0))
                ws.cell(row=row, column=8, value=p.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"packing_entries_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export dispatches
        dispatches = await db.dispatches_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if dispatches:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dispatches"
            headers = ["ID", "Challan Number", "Dispatch Type", "Dispatch Date", "Destination", "Products", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, d in enumerate(dispatches, 2):
                ws.cell(row=row, column=1, value=d.get("id", ""))
                ws.cell(row=row, column=2, value=d.get("challan_number", ""))
                ws.cell(row=row, column=3, value=d.get("dispatch_type", ""))
                ws.cell(row=row, column=4, value=d.get("date", ""))
                ws.cell(row=row, column=5, value=d.get("destination", ""))
                ws.cell(row=row, column=6, value=json.dumps(d.get("products", [])))
                ws.cell(row=row, column=7, value=d.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"dispatches_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export receives
        receives = await db.finished_product_receives_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if receives:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Receives"
            headers = ["ID", "SKU", "Quantity", "Receive Date", "Source", "Cost Per Unit", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, r in enumerate(receives, 2):
                ws.cell(row=row, column=1, value=r.get("id", ""))
                ws.cell(row=row, column=2, value=r.get("sku", ""))
                ws.cell(row=row, column=3, value=r.get("quantity", 0))
                ws.cell(row=row, column=4, value=r.get("receive_date", ""))
                ws.cell(row=row, column=5, value=r.get("source_name", ""))
                ws.cell(row=row, column=6, value=r.get("cost_per_unit", 0))
                ws.cell(row=row, column=7, value=r.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"receives_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export repacks
        repacks = await db.finished_product_repacks_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if repacks:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Repacks"
            headers = ["ID", "Source SKU", "Target SKU", "Quantity Used", "Quantity Produced", "Repack Date", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, rp in enumerate(repacks, 2):
                ws.cell(row=row, column=1, value=rp.get("id", ""))
                ws.cell(row=row, column=2, value=rp.get("source_sku", ""))
                ws.cell(row=row, column=3, value=rp.get("target_sku", ""))
                ws.cell(row=row, column=4, value=rp.get("quantity_used", 0))
                ws.cell(row=row, column=5, value=rp.get("quantity_produced", 0))
                ws.cell(row=row, column=6, value=rp.get("date", ""))
                ws.cell(row=row, column=7, value=rp.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"repacks_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export wastages
        wastages = await db.finished_product_wastages_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if wastages:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Wastages"
            headers = ["ID", "SKU", "Quantity", "Wastage Date", "Reason", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, w in enumerate(wastages, 2):
                ws.cell(row=row, column=1, value=w.get("id", ""))
                ws.cell(row=row, column=2, value=w.get("sku", ""))
                ws.cell(row=row, column=3, value=w.get("quantity", 0))
                ws.cell(row=row, column=4, value=w.get("date", ""))
                ws.cell(row=row, column=5, value=w.get("reason", ""))
                ws.cell(row=row, column=6, value=w.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"wastages_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export milk entries
        milk_entries = await db.milk_stock_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if milk_entries:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Milk Entries"
            headers = ["ID", "Date", "Supplier", "Quantity (kg)", "Fat %", "SNF %", "Fat (kg)", "SNF (kg)", "Total Amount", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, m in enumerate(milk_entries, 2):
                ws.cell(row=row, column=1, value=m.get("id", ""))
                ws.cell(row=row, column=2, value=m.get("date", ""))
                ws.cell(row=row, column=3, value=m.get("supplier", ""))
                ws.cell(row=row, column=4, value=m.get("quantity_kg", 0))
                ws.cell(row=row, column=5, value=m.get("fat_percent", 0))
                ws.cell(row=row, column=6, value=m.get("snf_percent", 0))
                ws.cell(row=row, column=7, value=m.get("fat_kg", 0))
                ws.cell(row=row, column=8, value=m.get("snf_kg", 0))
                ws.cell(row=row, column=9, value=m.get("total_amount", 0))
                ws.cell(row=row, column=10, value=m.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"milk_entries_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export raw material stocks
        rm_stocks = await db.raw_material_stock_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if rm_stocks:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Raw Material Purchases"
            headers = ["ID", "Material Name", "Quantity", "Unit", "Purchase Date", "Rate Per Unit", "Total Cost", "Supplier", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, rm in enumerate(rm_stocks, 2):
                ws.cell(row=row, column=1, value=rm.get("id", ""))
                ws.cell(row=row, column=2, value=rm.get("name", ""))
                ws.cell(row=row, column=3, value=rm.get("quantity", 0))
                ws.cell(row=row, column=4, value=rm.get("unit", ""))
                ws.cell(row=row, column=5, value=rm.get("purchase_date", ""))
                ws.cell(row=row, column=6, value=rm.get("rate_per_unit", 0))
                ws.cell(row=row, column=7, value=rm.get("total_cost", 0))
                ws.cell(row=row, column=8, value=rm.get("supplier", ""))
                ws.cell(row=row, column=9, value=rm.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"raw_material_purchases_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export raw material adjustments
        rm_adjs = await db.raw_material_adjustments_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if rm_adjs:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Raw Material Adjustments"
            headers = ["ID", "Material Name", "Adjustment Type", "Quantity", "Adjustment Date", "Reason", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, adj in enumerate(rm_adjs, 2):
                ws.cell(row=row, column=1, value=adj.get("id", ""))
                ws.cell(row=row, column=2, value=adj.get("material_name", ""))
                ws.cell(row=row, column=3, value=adj.get("adjustment_type", ""))
                ws.cell(row=row, column=4, value=adj.get("quantity", 0))
                ws.cell(row=row, column=5, value=adj.get("adjustment_date", ""))
                ws.cell(row=row, column=6, value=adj.get("reason", ""))
                ws.cell(row=row, column=7, value=adj.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"raw_material_adjustments_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Export raw material consumptions
        rm_cons = await db.raw_material_consumptions_archive.find({"archive_id": archive_id}, {"_id": 0}).to_list(100000)
        if rm_cons:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Raw Material Consumptions"
            headers = ["ID", "Material Name", "Quantity", "Consumption Date", "Purpose", "Notes"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row, con in enumerate(rm_cons, 2):
                ws.cell(row=row, column=1, value=con.get("id", ""))
                ws.cell(row=row, column=2, value=con.get("material_name", ""))
                ws.cell(row=row, column=3, value=con.get("quantity", 0))
                ws.cell(row=row, column=4, value=con.get("consumption_date", ""))
                ws.cell(row=row, column=5, value=con.get("purpose", ""))
                ws.cell(row=row, column=6, value=con.get("notes", ""))
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(f"raw_material_consumptions_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
        
        # Add closing stock summary
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Closing Stock Summary"
        ws.cell(row=1, column=1, value="Archive Date")
        ws.cell(row=1, column=2, value=archive['archive_date'])
        
        row = 3
        ws.cell(row=row, column=1, value="SEMI-FINISHED PRODUCTS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Product")
        ws.cell(row=row, column=2, value="Closing Stock (kg)")
        row += 1
        for name, stock in archive.get("closing_stock", {}).get("semi_finished_stock", {}).items():
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=round(stock, 2))
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="FINISHED PRODUCTS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="SKU")
        ws.cell(row=row, column=2, value="Closing Stock")
        row += 1
        for sku, stock in archive.get("closing_stock", {}).get("finished_stock", {}).items():
            ws.cell(row=row, column=1, value=sku)
            ws.cell(row=row, column=2, value=round(stock, 2))
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="RAW MATERIALS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Material")
        ws.cell(row=row, column=2, value="Closing Stock")
        row += 1
        for name, stock in archive.get("closing_stock", {}).get("raw_material_stock", {}).items():
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=round(stock, 2))
            row += 1
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        zip_file.writestr(f"closing_stock_summary_{archive['archive_date']}.xlsx", excel_buffer.getvalue())
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=archive_{archive['archive_date']}.zip"}
    )


@router.delete("/archive/{archive_id}")
async def delete_archive(
    archive_id: str,
    current_user: dict = Depends(require_admin)
):
    """Permanently delete an archive"""
    
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    # Delete from all archive collections
    await db.batches_archive.delete_many({"archive_id": archive_id})
    await db.packing_entries_archive.delete_many({"archive_id": archive_id})
    await db.dispatches_archive.delete_many({"archive_id": archive_id})
    await db.finished_product_receives_archive.delete_many({"archive_id": archive_id})
    await db.finished_product_repacks_archive.delete_many({"archive_id": archive_id})
    await db.finished_product_wastages_archive.delete_many({"archive_id": archive_id})
    await db.milk_stock_archive.delete_many({"archive_id": archive_id})
    await db.raw_material_stock_archive.delete_many({"archive_id": archive_id})
    await db.raw_material_adjustments_archive.delete_many({"archive_id": archive_id})
    await db.raw_material_consumptions_archive.delete_many({"archive_id": archive_id})
    
    # Delete metadata
    await db.archive_metadata.delete_one({"archive_id": archive_id})
    
    return {"success": True, "message": f"Archive {archive_id} deleted permanently"}


def calculate_batch_total_cost(batch: dict) -> dict:
    """Calculate and add cost fields to a batch document"""
    milk_kg = batch.get('milk_kg', 0) or 0
    fat_pct = batch.get('fat_percent', 0) or 0
    snf_pct = batch.get('snf_percent', 0) or 0
    fat_rate = batch.get('fat_rate', 0) or 0
    snf_rate = batch.get('snf_rate', 0) or 0
    qty = batch.get('quantity_produced', 0) or 0
    
    # Calculate milk cost
    fat_cost = round((fat_pct * 10 * milk_kg / 1000) * fat_rate, 2) if milk_kg > 0 else 0
    snf_cost = round((snf_pct * 10 * milk_kg / 1000) * snf_rate, 2) if milk_kg > 0 else 0
    milk_cost = round(fat_cost + snf_cost, 2)
    
    # Calculate other raw material cost
    other_rm_cost = 0
    for rm in batch.get('raw_materials', []) or []:
        other_rm_cost += (rm.get('quantity', 0) or 0) * (rm.get('cost_per_unit', 0) or 0)
    other_rm_cost = round(other_rm_cost, 2)
    
    # Calculate additional costs
    additional_costs_total = 0
    for c in batch.get('additional_costs', []) or []:
        additional_costs_total += c.get('amount', 0) or 0
    additional_costs_total = round(additional_costs_total, 2)
    
    # Total cost
    total_cost = round(milk_cost + other_rm_cost + additional_costs_total, 2)
    cost_per_unit = round(total_cost / qty, 2) if qty > 0 else 0
    
    # Add calculated fields to batch
    batch['fat_cost'] = fat_cost
    batch['snf_cost'] = snf_cost
    batch['milk_cost'] = milk_cost
    batch['other_rm_cost'] = other_rm_cost
    batch['additional_costs_total'] = additional_costs_total
    batch['total_cost'] = total_cost
    batch['cost_per_unit'] = cost_per_unit
    
    return batch


async def _build_closing_stock_summary(archive_id: str, archive: dict):
    """Build a per-SKU closing stock breakdown from archived data."""
    # Gather all data from archive collections
    packings = await db.packing_entries_archive.find(
        {"archive_id": archive_id}, {"_id": 0, "sku": 1, "quantity": 1, "source": 1}
    ).to_list(100000)
    receives = await db.finished_product_receives_archive.find(
        {"archive_id": archive_id}, {"_id": 0, "sku": 1, "quantity": 1}
    ).to_list(100000)
    dispatches = await db.dispatches_archive.find(
        {"archive_id": archive_id}, {"_id": 0, "products": 1}
    ).to_list(100000)
    repacks = await db.finished_product_repacks_archive.find(
        {"archive_id": archive_id}, {"_id": 0, "source_sku": 1, "target_sku": 1, "quantity_used": 1, "quantity_produced": 1}
    ).to_list(100000)
    wastages = await db.finished_product_wastages_archive.find(
        {"archive_id": archive_id}, {"_id": 0, "sku": 1, "quantity": 1}
    ).to_list(100000)

    # Aggregate per SKU
    sku_data = {}

    def get_sku(name):
        if name not in sku_data:
            sku_data[name] = {"produced": 0, "received": 0, "dispatched": 0, "repacked_out": 0, "wastage": 0, "type": "Finished"}
        return sku_data[name]

    for pk in packings:
        sku = pk.get('sku', '')
        qty = pk.get('quantity', 0) or 0
        source = pk.get('source')
        if source == 'receive':
            get_sku(sku)["received"] += qty
        else:
            get_sku(sku)["produced"] += qty

    for r in receives:
        sku = r.get('sku', '')
        qty = r.get('quantity', 0) or 0
        # Only count if not already counted from packing_entries_archive (avoid double count)
        # receives_archive is a separate collection, so count it
        # But the receive entries in packing_entries_archive are already counted above
        # So we skip here to avoid double-counting
        pass

    for d in dispatches:
        for p in d.get('products', []):
            sku = p.get('sku', '')
            qty = p.get('quantity', 0) or 0
            get_sku(sku)["dispatched"] += qty

    for rp in repacks:
        get_sku(rp.get('source_sku', ''))["repacked_out"] += rp.get('quantity_used', 0) or 0
        get_sku(rp.get('target_sku', ''))["produced"] += rp.get('quantity_produced', 0) or 0

    for w in wastages:
        get_sku(w.get('sku', ''))["wastage"] += w.get('quantity', 0) or 0

    # Also include semi-finished closing stock from archive
    closing = archive.get('closing_stock', {})
    sf_stock = closing.get('semi_finished_stock', {})
    for name, stock in sf_stock.items():
        if name not in sku_data:
            sku_data[name] = {"produced": 0, "received": 0, "dispatched": 0, "repacked_out": 0, "wastage": 0, "type": "Semi-Finished"}
        sku_data[name]["type"] = "Semi-Finished"
        sku_data[name]["closing_stock_override"] = round(stock, 2)

    result = []
    for name, d in sorted(sku_data.items()):
        if "closing_stock_override" in d:
            closing_stock = d["closing_stock_override"]
        else:
            closing_stock = round(d["produced"] + d["received"] - d["dispatched"] - d["repacked_out"] - d["wastage"], 2)
        result.append({
            "name": name,
            "type": d["type"],
            "produced": round(d["produced"], 2),
            "received": round(d["received"], 2),
            "dispatched": round(d["dispatched"], 2),
            "repacked_out": round(d["repacked_out"], 2),
            "wastage": round(d["wastage"], 2),
            "closing_stock": closing_stock
        })

    return {
        "archive_id": archive_id,
        "archive_date": archive["archive_date"],
        "collection": "closing_stock_summary",
        "data": result,
        "pagination": {"page": 1, "page_size": len(result), "total_count": len(result), "total_pages": 1}
    }



@router.get("/archive/{archive_id}/view")
async def view_archive(
    archive_id: str,
    collection: str = "batches",
    page: int = 1,
    page_size: int = 20,
    current_user: dict = Depends(require_admin)
):
    """View archived data by collection with pagination"""
    
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    # Map collection names to archive collections
    collection_map = {
        "batches": db.batches_archive,
        "packing_entries": db.packing_entries_archive,
        "dispatches": db.dispatches_archive,
        "receives": db.finished_product_receives_archive,
        "repacks": db.finished_product_repacks_archive,
        "wastages": db.finished_product_wastages_archive,
        "milk_entries": db.milk_stock_archive,
        "raw_material_stocks": db.raw_material_stock_archive,
        "raw_material_adjustments": db.raw_material_adjustments_archive,
        "raw_material_consumptions": db.raw_material_consumptions_archive
    }
    
    if collection not in collection_map and collection != "closing_stock_summary":
        raise HTTPException(status_code=400, detail=f"Invalid collection. Valid options: {list(collection_map.keys()) + ['closing_stock_summary']}")
    
    # Special handling for closing stock summary
    if collection == "closing_stock_summary":
        return await _build_closing_stock_summary(archive_id, archive)
    
    archive_collection = collection_map[collection]
    
    # Get total count
    total_count = await archive_collection.count_documents({"archive_id": archive_id})
    total_pages = (total_count + page_size - 1) // page_size
    
    # Get paginated data
    skip = (page - 1) * page_size
    
    # For packing_entries, exclude receives (they show in Receives tab)
    query_filter = {"archive_id": archive_id}
    if collection == "packing_entries":
        query_filter["source"] = {"$nin": ["receive"]}
    
    # Recalculate total count with filter
    if collection == "packing_entries":
        total_count = await archive_collection.count_documents(query_filter)
        total_pages = (total_count + page_size - 1) // page_size
    
    data = await archive_collection.find(
        query_filter, 
        {"_id": 0}
    ).skip(skip).limit(page_size).to_list(page_size)
    
    # For batches collection, calculate total costs and get remaining stock
    if collection == "batches":
        batch_ids = [b.get('id') for b in data if b.get('id')]
        
        # Check live semi_finished_products first
        sf_records = await db.semi_finished_products.find(
            {"batch_id": {"$in": batch_ids}},
            {"_id": 0, "batch_id": 1, "current_stock": 1}
        ).to_list(1000)
        stock_map = {sf['batch_id']: sf.get('current_stock', 0) for sf in sf_records}
        
        # Check live finished_products for direct batch→finished
        fp_records = await db.finished_products.find(
            {"batch_id": {"$in": batch_ids}, "source": "batch"},
            {"_id": 0, "batch_id": 1, "current_stock": 1}
        ).to_list(1000)
        for fp in fp_records:
            if fp['batch_id'] not in stock_map:
                stock_map[fp['batch_id']] = fp.get('current_stock', 0)
        
        # For batches not found in live collections, calculate from archived data
        missing_ids = [bid for bid in batch_ids if bid not in stock_map]
        if missing_ids:
            # Get archived packing entries to calculate consumption per batch
            archived_packings = await db.packing_entries_archive.find(
                {"archive_id": archive_id, "batch_id": {"$in": missing_ids}},
                {"_id": 0, "batch_id": 1, "semi_finished_consumed": 1, "quantity": 1, "quantity_wasted": 1}
            ).to_list(10000)
            
            consumed_per_batch = {}
            for pk in archived_packings:
                bid = pk.get('batch_id', '')
                consumed = pk.get('semi_finished_consumed') or (pk.get('quantity', 0) + pk.get('quantity_wasted', 0))
                consumed_per_batch[bid] = consumed_per_batch.get(bid, 0) + consumed
            
            for bid in missing_ids:
                batch_doc = next((b for b in data if b.get('id') == bid), None)
                if batch_doc:
                    produced = batch_doc.get('quantity_produced', 0)
                    consumed = consumed_per_batch.get(bid, 0)
                    stock_map[bid] = max(0, produced - consumed)
        
        # Calculate costs and add remaining stock
        for batch in data:
            calculate_batch_total_cost(batch)
            batch['remaining_stock'] = round(stock_map.get(batch.get('id'), 0), 2)
    
    # For packing_entries, add source label for batch-direct entries
    if collection == "packing_entries":
        for entry in data:
            if entry.get('source') == 'batch':
                entry['semi_finished_product'] = '(Direct from Batch)'
    
    return {
        "archive_id": archive_id,
        "archive_date": archive["archive_date"],
        "collection": collection,
        "data": data,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages
        }
    }




@router.get("/archive/alert-status")
async def get_archive_alert_status(current_user: dict = Depends(get_current_user)):
    """Check if archive alert should be shown"""
    if current_user.get("role") != "admin":
        return {"show_alert": False}
    
    setting = await db.settings.find_one({"key": "last_archive_date"})
    
    if not setting:
        # No archive ever done
        return {
            "show_alert": True,
            "message": "Data has never been archived. Consider archiving old data to maintain system performance.",
            "last_archive_date": None
        }
    
    last_archive = datetime.strptime(setting["value"], "%Y-%m-%d")
    months_since = (datetime.now() - last_archive).days / 30
    
    if months_since > 3:
        return {
            "show_alert": True,
            "message": f"Data not archived for over {int(months_since)} months. Consider archiving old data to maintain system performance.",
            "last_archive_date": setting["value"]
        }
    
    return {
        "show_alert": False,
        "last_archive_date": setting["value"]
    }


# Activity log cleanup (runs on startup and can be called manually)
async def cleanup_old_activity_logs():
    """Delete activity logs older than 30 days"""
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    result = await db.activity_logs.delete_many({"timestamp": {"$lt": cutoff_date}})
    return result.deleted_count


@router.post("/archive/cleanup-activity-logs")
async def manual_cleanup_activity_logs(current_user: dict = Depends(require_admin)):
    """Manually trigger activity log cleanup"""
    deleted_count = await cleanup_old_activity_logs()
    return {"success": True, "deleted_count": deleted_count}



@router.post("/archive/{archive_id}/backfill-packing")
async def backfill_packing_entries(archive_id: str, current_user: dict = Depends(require_admin)):
    """
    Backfill packing entries from finished_products into an existing archive.
    This is a one-time migration for archives created before the fix.
    """
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    cutoff_str = archive["archive_date"]
    
    # Get finished_products dated on or before archive date
    packings = await db.finished_products.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    
    if not packings:
        return {"success": True, "message": "No packing entries to backfill", "count": 0}
    
    # Check if already backfilled
    existing = await db.packing_entries_archive.count_documents({"archive_id": archive_id})
    if existing > 0:
        return {"success": False, "message": f"Archive already has {existing} packing entries. Delete them first if you want to re-backfill."}
    
    # Enrich packing entries with semi-finished product names
    for p in packings:
        p["archive_id"] = archive_id
        if p.get('semi_finished_id'):
            sf = await db.semi_finished_products.find_one({"id": p['semi_finished_id']}, {"_id": 0, "product_name": 1})
            p['semi_finished_product'] = sf.get('product_name', 'Unknown') if sf else 'Unknown'
        else:
            p['semi_finished_product'] = 'Unknown'
        p['finished_product_sku'] = p.get('sku', '')
        p['packing_date'] = p.get('date', '')
        p['quantity_produced'] = p.get('quantity', 0)
    
    await db.packing_entries_archive.insert_many(packings)
    
    return {"success": True, "message": f"Backfilled {len(packings)} packing entries", "count": len(packings)}


@router.post("/archive/{archive_id}/backfill-all")
async def backfill_all_collections(archive_id: str, current_user: dict = Depends(require_admin)):
    """
    Backfill ALL collections into an existing archive.
    This is a one-time migration for archives that missed some data.
    """
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    cutoff_str = archive["archive_date"]
    results = {}
    
    # Backfill dispatches
    dispatches = await db.dispatches.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_dispatches = await db.dispatches_archive.count_documents({"archive_id": archive_id})
    if dispatches and existing_dispatches == 0:
        for d in dispatches:
            d["archive_id"] = archive_id
        await db.dispatches_archive.insert_many(dispatches)
        results["dispatches"] = len(dispatches)
    else:
        results["dispatches"] = f"skipped ({existing_dispatches} existing)"
    
    # Backfill receives
    receives = await db.finished_product_receives.find({"receive_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_receives = await db.finished_product_receives_archive.count_documents({"archive_id": archive_id})
    if receives and existing_receives == 0:
        for r in receives:
            r["archive_id"] = archive_id
        await db.finished_product_receives_archive.insert_many(receives)
        results["receives"] = len(receives)
    else:
        results["receives"] = f"skipped ({existing_receives} existing)"
    
    # Backfill repacks
    repacks = await db.finished_product_repacks.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_repacks = await db.finished_product_repacks_archive.count_documents({"archive_id": archive_id})
    if repacks and existing_repacks == 0:
        for rp in repacks:
            rp["archive_id"] = archive_id
        await db.finished_product_repacks_archive.insert_many(repacks)
        results["repacks"] = len(repacks)
    else:
        results["repacks"] = f"skipped ({existing_repacks} existing)"
    
    # Backfill wastages
    wastages = await db.finished_product_wastages.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_wastages = await db.finished_product_wastages_archive.count_documents({"archive_id": archive_id})
    if wastages and existing_wastages == 0:
        for w in wastages:
            w["archive_id"] = archive_id
        await db.finished_product_wastages_archive.insert_many(wastages)
        results["wastages"] = len(wastages)
    else:
        results["wastages"] = f"skipped ({existing_wastages} existing)"
    
    # Backfill milk entries
    milk_entries = await db.milk_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_milk = await db.milk_stock_archive.count_documents({"archive_id": archive_id})
    if milk_entries and existing_milk == 0:
        for m in milk_entries:
            m["archive_id"] = archive_id
        await db.milk_stock_archive.insert_many(milk_entries)
        results["milk_entries"] = len(milk_entries)
    else:
        results["milk_entries"] = f"skipped ({existing_milk} existing)"
    
    # Backfill raw material stocks (purchases)
    rm_stocks = await db.raw_material_stock.find({"date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_rm_stocks = await db.raw_material_stock_archive.count_documents({"archive_id": archive_id})
    if rm_stocks and existing_rm_stocks == 0:
        for rs in rm_stocks:
            rs["archive_id"] = archive_id
        await db.raw_material_stock_archive.insert_many(rm_stocks)
        results["raw_material_stocks"] = len(rm_stocks)
    else:
        results["raw_material_stocks"] = f"skipped ({existing_rm_stocks} existing)"
    
    # Backfill raw material adjustments
    rm_adjs = await db.raw_material_adjustments.find({"adjustment_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_rm_adjs = await db.raw_material_adjustments_archive.count_documents({"archive_id": archive_id})
    if rm_adjs and existing_rm_adjs == 0:
        for ra in rm_adjs:
            ra["archive_id"] = archive_id
        await db.raw_material_adjustments_archive.insert_many(rm_adjs)
        results["raw_material_adjustments"] = len(rm_adjs)
    else:
        results["raw_material_adjustments"] = f"skipped ({existing_rm_adjs} existing)"
    
    # Backfill raw material consumptions
    rm_cons = await db.raw_material_consumptions.find({"consumption_date": {"$lte": cutoff_str}}, {"_id": 0}).to_list(100000)
    existing_rm_cons = await db.raw_material_consumptions_archive.count_documents({"archive_id": archive_id})
    if rm_cons and existing_rm_cons == 0:
        for rc in rm_cons:
            rc["archive_id"] = archive_id
        await db.raw_material_consumptions_archive.insert_many(rc)
        results["raw_material_consumptions"] = len(rm_cons)
    else:
        results["raw_material_consumptions"] = f"skipped ({existing_rm_cons} existing)"
    
    return {"success": True, "backfilled": results}


@router.post("/archive/fix-opening-balances")
async def fix_opening_balances(current_user: dict = Depends(require_admin)):
    """
    Fix missing finished_products entries for opening balance receives.
    This is needed for archives created before the fix.
    """
    # Find all opening balance receives that don't have corresponding finished_products entries
    ob_receives = await db.finished_product_receives.find({"is_opening_balance": True}, {"_id": 0}).to_list(1000)
    
    created_fp = 0
    for ob in ob_receives:
        receive_id = ob['id']
        sku = ob['sku']
        
        # Check if finished_products entry exists
        existing = await db.finished_products.find_one({"source_receive_id": receive_id}, {"_id": 0})
        if not existing:
            # Create the missing entry
            await db.finished_products.insert_one({
                "id": f"FP-{receive_id}",
                "sku": sku,
                "quantity": ob['quantity'],
                "quantity_wasted": 0,
                "current_stock": ob['quantity'],
                "unit": ob.get('unit', ''),
                "source": "receive",
                "source_receive_id": receive_id,
                "date": ob['receive_date'],
                "is_opening_balance": True,
                "notes": ob.get('notes', ''),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            created_fp += 1
    
    # Also fix semi-finished products opening balances
    ob_batches = await db.batches.find({"is_opening_balance": True, "output_type": "semi-finished"}, {"_id": 0}).to_list(1000)
    
    created_sf = 0
    for ob in ob_batches:
        batch_id = ob['id']
        product_name = ob['product_name']
        
        # Check if semi_finished_products entry exists
        existing = await db.semi_finished_products.find_one({"batch_id": batch_id}, {"_id": 0})
        if not existing:
            # Create the missing entry
            await db.semi_finished_products.insert_one({
                "id": f"SF-{batch_id}",
                "product_name": product_name,
                "batch_id": batch_id,
                "quantity_kg": ob['quantity_produced'],
                "current_stock": ob['quantity_produced'],
                "date": ob['date'],
                "is_opening_balance": True,
                "notes": ob.get('notes', ''),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            created_sf += 1
    
    return {
        "success": True, 
        "finished_products_created": created_fp,
        "semi_finished_created": created_sf,
        "message": f"Created {created_fp} finished_products and {created_sf} semi_finished_products entries"
    }


@router.post("/archive/{archive_id}/cleanup-old-data")
async def cleanup_old_data(archive_id: str, current_user: dict = Depends(require_admin)):
    """
    Clean up old data that should have been deleted during archive.
    This is a one-time fix for archives created before the delete fix.
    WARNING: This will delete old records from active collections.
    """
    # Verify archive exists
    archive = await db.archive_metadata.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    cutoff_str = archive["archive_date"]
    results = {}
    
    # Delete old finished_products (but keep opening balance entries)
    old_fp = await db.finished_products.count_documents({
        "date": {"$lte": cutoff_str},
        "$or": [
            {"is_opening_balance": {"$exists": False}},
            {"is_opening_balance": False}
        ]
    })
    if old_fp > 0:
        await db.finished_products.delete_many({
            "date": {"$lte": cutoff_str},
            "$or": [
                {"is_opening_balance": {"$exists": False}},
                {"is_opening_balance": False}
            ]
        })
    results["finished_products_deleted"] = old_fp
    
    # Delete old semi_finished_products (but keep opening balance entries)
    old_sf = await db.semi_finished_products.count_documents({
        "date": {"$lte": cutoff_str},
        "$or": [
            {"is_opening_balance": {"$exists": False}},
            {"is_opening_balance": False}
        ]
    })
    if old_sf > 0:
        await db.semi_finished_products.delete_many({
            "date": {"$lte": cutoff_str},
            "$or": [
                {"is_opening_balance": {"$exists": False}},
                {"is_opening_balance": False}
            ]
        })
    results["semi_finished_deleted"] = old_sf
    
    # Delete old dispatches
    old_dispatches = await db.dispatches.count_documents({"date": {"$lte": cutoff_str}})
    if old_dispatches > 0:
        await db.dispatches.delete_many({"date": {"$lte": cutoff_str}})
    results["dispatches_deleted"] = old_dispatches
    
    # Delete old repacks
    old_repacks = await db.finished_product_repacks.count_documents({"date": {"$lte": cutoff_str}})
    if old_repacks > 0:
        await db.finished_product_repacks.delete_many({"date": {"$lte": cutoff_str}})
    results["repacks_deleted"] = old_repacks
    
    # Delete old wastages
    old_wastages = await db.finished_product_wastages.count_documents({"date": {"$lte": cutoff_str}})
    if old_wastages > 0:
        await db.finished_product_wastages.delete_many({"date": {"$lte": cutoff_str}})
    results["wastages_deleted"] = old_wastages
    
    # Delete old receives (but keep opening balance entries)
    old_receives = await db.finished_product_receives.count_documents({
        "receive_date": {"$lte": cutoff_str},
        "$or": [
            {"is_opening_balance": {"$exists": False}},
            {"is_opening_balance": False}
        ]
    })
    if old_receives > 0:
        await db.finished_product_receives.delete_many({
            "receive_date": {"$lte": cutoff_str},
            "$or": [
                {"is_opening_balance": {"$exists": False}},
                {"is_opening_balance": False}
            ]
        })
    results["receives_deleted"] = old_receives
    
    # Delete old raw material stocks (but keep opening balance entries)
    old_rm_stocks = await db.raw_material_stock.count_documents({
        "date": {"$lte": cutoff_str},
        "$or": [
            {"is_opening_balance": {"$exists": False}},
            {"is_opening_balance": False}
        ]
    })
    if old_rm_stocks > 0:
        await db.raw_material_stock.delete_many({
            "date": {"$lte": cutoff_str},
            "$or": [
                {"is_opening_balance": {"$exists": False}},
                {"is_opening_balance": False}
            ]
        })
    results["raw_material_stocks_deleted"] = old_rm_stocks
    
    # Delete old raw material adjustments
    old_rm_adjs = await db.rm_adjustments.count_documents({"date": {"$lte": cutoff_str}})
    if old_rm_adjs > 0:
        await db.rm_adjustments.delete_many({"date": {"$lte": cutoff_str}})
    results["raw_material_adjustments_deleted"] = old_rm_adjs
    
    # Delete old raw material consumptions
    old_rm_cons = await db.rm_direct_consumption.count_documents({"date": {"$lte": cutoff_str}})
    if old_rm_cons > 0:
        await db.rm_direct_consumption.delete_many({"date": {"$lte": cutoff_str}})
    results["raw_material_consumptions_deleted"] = old_rm_cons
    
    return {
        "success": True,
        "archive_date": cutoff_str,
        "results": results,
        "message": f"Cleanup complete for data dated <= {cutoff_str}"
    }










@router.get("/archive/recalculate-stock")
async def recalculate_stock(current_user: dict = Depends(require_admin)):
    """
    Recalculate current stock from ALL transactions.
    This is useful to verify stock integrity or find discrepancies.
    """
    
    # Calculate semi-finished stock
    semi_finished_stock = {}
    
    # From batches (production)
    batches = await db.batches.find({}, {"_id": 0}).to_list(100000)
    for batch in batches:
        if batch.get("output_type") == "semi-finished":
            name = batch.get("product_name", "Unknown")
            qty = float(batch.get("quantity_produced", 0))
            semi_finished_stock[name] = semi_finished_stock.get(name, 0) + qty
    
    # From packing/finished_products (consumption of semi-finished)
    packings = await db.finished_products.find({}, {"_id": 0}).to_list(100000)
    for pk in packings:
        sf_id = pk.get("semi_finished_id")
        if sf_id:
            sf = await db.semi_finished_products.find_one({"id": sf_id}, {"_id": 0, "product_name": 1})
            name = sf.get("product_name", "Unknown") if sf else "Unknown"
        else:
            name = "Unknown"
        consumed = float(pk.get("semi_finished_consumed", 0))
        semi_finished_stock[name] = semi_finished_stock.get(name, 0) - consumed
    
    # Calculate finished product stock
    finished_stock = {}
    
    # From batches (direct finished)
    for batch in batches:
        if batch.get("output_type") == "finished":
            sku = batch.get("product_name", "Unknown")
            qty = float(batch.get("quantity_produced", 0))
            finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From packing/finished_products (production of finished goods)
    for pk in packings:
        sku = pk.get("sku", "Unknown")
        qty = float(pk.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From receives
    receives = await db.finished_product_receives.find({}, {"_id": 0}).to_list(100000)
    for r in receives:
        sku = r.get("sku", "Unknown")
        qty = float(r.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) + qty
    
    # From repacks
    repacks = await db.finished_product_repacks.find({}, {"_id": 0}).to_list(100000)
    for rp in repacks:
        source = rp.get("source_sku", "Unknown")
        target = rp.get("target_sku", "Unknown")
        used = float(rp.get("quantity_used", 0))
        produced = float(rp.get("quantity_produced", 0))
        finished_stock[source] = finished_stock.get(source, 0) - used
        finished_stock[target] = finished_stock.get(target, 0) + produced
    
    # From dispatches
    dispatches = await db.dispatches.find({}, {"_id": 0}).to_list(100000)
    for d in dispatches:
        for prod in d.get("products", []):
            sku = prod.get("sku", "Unknown")
            qty = float(prod.get("quantity", 0))
            finished_stock[sku] = finished_stock.get(sku, 0) - qty
    
    # From wastages
    wastages = await db.finished_product_wastages.find({}, {"_id": 0}).to_list(100000)
    for w in wastages:
        sku = w.get("sku", "Unknown")
        qty = float(w.get("quantity", 0))
        finished_stock[sku] = finished_stock.get(sku, 0) - qty
    
    # Calculate raw material stock
    raw_material_stock = {}
    
    # From initial stocks
    rm_initials = await db.initial_stocks.find({"type": "raw_material"}, {"_id": 0}).to_list(100)
    for init in rm_initials:
        name = init.get("name", "Unknown")
        qty = float(init.get("quantity", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) + qty
    
    # From daily stock entries (contains purchased and used)
    rm_stocks = await db.raw_material_stock.find({}, {"_id": 0}).to_list(100000)
    for rm in rm_stocks:
        name = rm.get("name", "Unknown")
        purchased = float(rm.get("purchased", 0))
        used = float(rm.get("used", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) + purchased - used
    
    # From adjustments
    rm_adjs = await db.rm_adjustments.find({}, {"_id": 0}).to_list(100000)
    for adj in rm_adjs:
        name = adj.get("material_name", adj.get("name", "Unknown"))
        qty = float(adj.get("quantity", 0))
        adj_type = adj.get("adjustment_type", adj.get("type", "gain"))
        if adj_type == "loss":
            qty = -qty
        raw_material_stock[name] = raw_material_stock.get(name, 0) + qty
    
    # From direct consumptions
    rm_cons = await db.rm_direct_consumption.find({}, {"_id": 0}).to_list(100000)
    for con in rm_cons:
        name = con.get("material_name", con.get("name", "Unknown"))
        qty = float(con.get("quantity", 0))
        raw_material_stock[name] = raw_material_stock.get(name, 0) - qty
    
    # Calculate milk stock (including Fat and SNF)
    milk_qty = 0
    milk_fat = 0
    milk_snf = 0
    
    # From initial stocks
    milk_initials = await db.initial_stocks.find({"type": "milk"}, {"_id": 0}).to_list(100)
    for init in milk_initials:
        milk_qty += float(init.get("quantity", 0))
        milk_fat += float(init.get("fat_kg", 0))
        milk_snf += float(init.get("snf_kg", 0))
    
    # From milk_stock entries (received/purchased milk)
    milk_stock_entries = await db.milk_stock.find({}, {"_id": 0}).to_list(100000)
    for entry in milk_stock_entries:
        milk_qty += float(entry.get("quantity_kg", 0))
        milk_fat += float(entry.get("fat_kg", 0))
        milk_snf += float(entry.get("snf_kg", 0))
    
    # From batches (milk consumed in production)
    for batch in batches:
        batch_milk_kg = float(batch.get("milk_kg", 0))
        if batch_milk_kg > 0:
            milk_qty -= batch_milk_kg
            fat_pct = float(batch.get("fat_percent", 0))
            snf_pct = float(batch.get("snf_percent", 0))
            milk_fat -= (fat_pct / 100) * batch_milk_kg
            milk_snf -= (snf_pct / 100) * batch_milk_kg
    
    # From milk wastages
    milk_wastages = await db.milk_wastages.find({}, {"_id": 0}).to_list(100000)
    for w in milk_wastages:
        milk_qty -= float(w.get("quantity_kg", 0))
        milk_fat -= float(w.get("fat_kg", 0))
        milk_snf -= float(w.get("snf_kg", 0))
    
    # From milk adjustments
    milk_adjs = await db.milk_adjustments.find({}, {"_id": 0}).to_list(100000)
    for adj in milk_adjs:
        qty = float(adj.get("quantity_kg", 0))
        fat = float(adj.get("fat_kg", 0))
        snf = float(adj.get("snf_kg", 0))
        adj_type = adj.get("type", "gain")
        if adj_type == "loss":
            qty, fat, snf = -qty, -fat, -snf
        milk_qty += qty
        milk_fat += fat
        milk_snf += snf
    
    milk_stock = {
        "Milk": round(milk_qty, 2),
        "Fat": round(milk_fat, 2),
        "SNF": round(milk_snf, 2)
    }
    
    return {
        "calculated_at": datetime.now(timezone.utc).isoformat(),
        "semi_finished_stock": {k: round(v, 2) for k, v in semi_finished_stock.items()},
        "finished_stock": {k: round(v, 2) for k, v in finished_stock.items()},
        "raw_material_stock": {k: round(v, 2) for k, v in raw_material_stock.items()},
        "milk_stock": {k: round(v, 2) for k, v in milk_stock.items()},
        "note": "This is the calculated stock from all transactions. Compare with displayed stock to find discrepancies."
    }
